"""Build the Excel executive report from the Power BI-ready dataset.

Usage:
    python src/export_excel.py
    python src/export_excel.py --output /path/to/report.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "powerbi" / "luas_rent_powerbi.csv"
DEFAULT_OUTPUT = ROOT / "excel" / "executive_report.xlsx"

GREEN = "00A651"
DARK = "1C3F5F"
GOLD = "F2A900"
RED = "C0392B"
LIGHT_GREEN = "E8F5EE"
LIGHT_BLUE = "EAF0F5"
LIGHT_GREY = "F3F6F5"
MID_GREY = "D9E1E3"
WHITE = "FFFFFF"
BLACK = "1C1C1C"

THIN_GREY = Side(style="thin", color=MID_GREY)


def style_title(ws, cell_range: str, text: str, size: int = 22) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.font = Font(name="Aptos Display", size=size, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(vertical="center")


def style_subtitle(ws, cell_range: str, text: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.font = Font(name="Aptos", size=11, italic=True, color=DARK)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_common_sheet_style(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def load_data() -> pd.DataFrame:
    if not SOURCE.exists():
        raise FileNotFoundError(
            f"Missing {SOURCE}. Run 'python src/export_powerbi.py' first."
        )
    df = pd.read_csv(SOURCE)
    expected = {
        "stop_sequence", "station", "travel_min_to_centre", "avg_rent",
        "dublin_avg", "year", "bedrooms", "property_type", "rtb_area",
    }
    missing = expected - set(df.columns)
    if missing:
        raise ValueError(f"Power BI export is missing columns: {sorted(missing)}")
    return df


def latest_station_slice(df: pd.DataFrame) -> pd.DataFrame:
    latest = int(df["year"].max())
    out = df[
        (df["year"] == latest)
        & (df["bedrooms"] == "1 to 2 bed")
        & (df["property_type"] == "All property types")
    ].copy()
    out = out.sort_values("stop_sequence").drop_duplicates("station")
    if len(out) != 24:
        raise ValueError(f"Expected 24 stations in latest slice, got {len(out)}")
    return out


def yearly_trend(df: pd.DataFrame) -> pd.DataFrame:
    selected = df[
        (df["bedrooms"] == "1 to 2 bed")
        & (df["property_type"] == "All property types")
    ].copy()
    rows = []
    for year, group in selected.groupby("year", sort=True):
        x = group["travel_min_to_centre"].astype(float)
        y = group["avg_rent"].astype(float)
        slope = x.cov(y) / x.var(ddof=1)
        intercept = y.mean() - slope * x.mean()
        fitted = intercept + slope * x
        ss_res = ((y - fitted) ** 2).sum()
        ss_tot = ((y - y.mean()) ** 2).sum()
        r2 = 1 - ss_res / ss_tot if ss_tot else 0
        rows.append({
            "year": int(year),
            "line_avg": group["avg_rent"].mean(),
            "dublin_avg": group["dublin_avg"].iloc[0],
            "gradient": slope,
            "r_squared": r2,
        })
    return pd.DataFrame(rows)


def build_station_analysis(wb: Workbook, station_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Station Analysis")
    set_common_sheet_style(ws)
    style_title(ws, "A1:L1", "Station Analysis — 2025")
    style_subtitle(
        ws,
        "A2:L2",
        "Like-for-like comparison: RTB 1 to 2 bed, all property types. "
        "Savings are measured against St. Stephen's Green.",
    )
    headers = [
        "Sequence", "Station", "Commute (min)", "Avg Rent (€)",
        "Dublin Avg (€)", "Vs Dublin", "Monthly Saving (€)",
        "Saving / Extra Min (€)", "Change vs Previous Stop (€)",
        "Rent Rank", "RTB Area", "Park & Ride",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_header_row(ws, 4, 1, len(headers))

    first_data_row = 5
    last_data_row = first_data_row + len(station_df) - 1
    for idx, (_, row) in enumerate(station_df.iterrows(), first_data_row):
        ws.cell(idx, 1, int(row["stop_sequence"]))
        ws.cell(idx, 2, row["station"])
        ws.cell(idx, 3, int(row["travel_min_to_centre"]))
        ws.cell(idx, 4, float(row["avg_rent"]))
        ws.cell(idx, 5, float(row["dublin_avg"]))
        ws.cell(idx, 6, f'=IFERROR((D{idx}-E{idx})/E{idx},"")')
        ws.cell(idx, 7, f"=$D${first_data_row}-D{idx}")
        ws.cell(idx, 8, f'=IF(C{idx}=0,"",G{idx}/C{idx})')
        ws.cell(idx, 9, "" if idx == first_data_row else f"=D{idx}-D{idx-1}")
        ws.cell(idx, 10, f"=RANK(D{idx},$D${first_data_row}:$D${last_data_row},1)")
        ws.cell(idx, 11, row["rtb_area"])
        ws.cell(idx, 12, row["park_and_ride"])

    add_table(ws, f"A4:L{last_data_row}", "StationAnalysisTable")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{last_data_row}"
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 34
    widths = {
        "A": 10, "B": 24, "C": 14, "D": 15, "E": 15, "F": 13,
        "G": 18, "H": 22, "I": 26, "J": 11, "K": 30, "L": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(first_data_row, last_data_row + 1):
        for col in range(1, 13):
            ws.cell(row, col).alignment = Alignment(vertical="center")
        for col in (4, 5, 7, 8, 9):
            ws.cell(row, col).number_format = '€#,##0'
        ws.cell(row, 6).number_format = "0.0%"

    ws.conditional_formatting.add(
        f"D{first_data_row}:D{last_data_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"H{first_data_row}:H{last_data_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:L{last_data_row}"


def build_trends(wb: Workbook, trend: pd.DataFrame) -> None:
    ws = wb.create_sheet("Trends")
    set_common_sheet_style(ws)
    style_title(ws, "A1:F1", "Rent Trends — 2008 to 2025")
    style_subtitle(
        ws,
        "A2:F2",
        "Green Line station average versus Dublin county, using the 1 to 2 bed / all property types series.",
    )
    headers = ["Year", "Green Line Avg (€)", "Dublin Avg (€)", "Premium vs Dublin", "Gradient €/min", "R²"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    style_header_row(ws, 4, 1, len(headers))
    first = 5
    for idx, (_, row) in enumerate(trend.iterrows(), first):
        ws.cell(idx, 1, int(row["year"]))
        ws.cell(idx, 2, float(row["line_avg"]))
        ws.cell(idx, 3, float(row["dublin_avg"]))
        ws.cell(idx, 4, f'=IFERROR((B{idx}-C{idx})/C{idx},"")')
        ws.cell(idx, 5, float(row["gradient"]))
        ws.cell(idx, 6, float(row["r_squared"]))
        for col in (2, 3, 5):
            ws.cell(idx, col).number_format = '€#,##0.0'
        for col in (4, 6):
            ws.cell(idx, col).number_format = "0.0%"
    last = first + len(trend) - 1
    add_table(ws, f"A4:F{last}", "TrendTable")
    ws.freeze_panes = "A5"
    for col, width in {"A": 10, "B": 20, "C": 18, "D": 20, "E": 18, "F": 12}.items():
        ws.column_dimensions[col].width = width
    ws.conditional_formatting.add(
        f"D{first}:D{last}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )

    chart = LineChart()
    chart.title = "Green Line vs Dublin County"
    chart.y_axis.title = "Average monthly rent (€)"
    chart.x_axis.title = "Year"
    chart.height = 8
    chart.width = 15
    chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=4, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last))
    chart.legend.position = "b"
    ws.add_chart(chart, "H4")

    gradient_chart = LineChart()
    gradient_chart.title = "Rent Gradient by Year"
    gradient_chart.y_axis.title = "€ per extra commute minute"
    gradient_chart.x_axis.title = "Year"
    gradient_chart.height = 8
    gradient_chart.width = 15
    gradient_chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=last), titles_from_data=True)
    gradient_chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last))
    gradient_chart.legend = None
    ws.add_chart(gradient_chart, "H20")
    ws.print_area = "A1:N35"


def build_pivot_summary(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Pivot Summary")
    set_common_sheet_style(ws)
    years = sorted(int(y) for y in df["year"].unique())
    selected = df[
        (df["bedrooms"] == "1 to 2 bed")
        & (df["property_type"] == "All property types")
    ]
    pivot = selected.pivot_table(index="station", columns="year", values="avg_rent", aggfunc="first")
    order = (
        selected[["station", "stop_sequence"]]
        .drop_duplicates()
        .sort_values("stop_sequence")["station"]
        .tolist()
    )
    pivot = pivot.reindex(order)

    end_col = 1 + len(years)
    from openpyxl.utils import get_column_letter
    end_letter = get_column_letter(end_col)
    style_title(ws, f"A1:{end_letter}1", "Pivot Summary")
    style_subtitle(
        ws,
        f"A2:{end_letter}2",
        "Average registered rent by station and year. Green cells are lower rents; red cells are higher rents.",
    )
    ws.cell(4, 1, "Station")
    for col, year in enumerate(years, 2):
        ws.cell(4, col, str(year))
    style_header_row(ws, 4, 1, end_col)
    first = 5
    for row_idx, station in enumerate(order, first):
        ws.cell(row_idx, 1, station)
        for col_idx, year in enumerate(years, 2):
            value = pivot.loc[station, year]
            ws.cell(row_idx, col_idx, None if pd.isna(value) else float(value))
            ws.cell(row_idx, col_idx).number_format = '€#,##0'
    last = first + len(order) - 1
    add_table(ws, f"A4:{end_letter}{last}", "StationYearPivotTable")
    ws.conditional_formatting.add(
        f"B{first}:{end_letter}{last}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )

    second_title_row = last + 3
    ws.cell(second_title_row, 1, "2025 rent by bedroom category")
    ws.cell(second_title_row, 1).font = Font(name="Aptos Display", size=15, bold=True, color=DARK)
    bed_order = ["One bed", "1 to 2 bed", "Two bed", "Three bed", "Four plus bed", "All bedrooms"]
    bedroom = df[
        (df["year"] == df["year"].max())
        & (df["property_type"] == "All property types")
        & (df["bedrooms"].isin(bed_order))
    ].pivot_table(index="station", columns="bedrooms", values="avg_rent", aggfunc="first")
    bedroom = bedroom.reindex(order)
    header_row = second_title_row + 2
    ws.cell(header_row, 1, "Station")
    for col, bed in enumerate(bed_order, 2):
        ws.cell(header_row, col, bed)
    style_header_row(ws, header_row, 1, 1 + len(bed_order))
    data_start = header_row + 1
    for row_idx, station in enumerate(order, data_start):
        ws.cell(row_idx, 1, station)
        for col_idx, bed in enumerate(bed_order, 2):
            value = bedroom.loc[station, bed] if bed in bedroom.columns else None
            ws.cell(row_idx, col_idx, None if value is None or pd.isna(value) else float(value))
            ws.cell(row_idx, col_idx).number_format = '€#,##0'
    data_end = data_start + len(order) - 1
    second_end_letter = get_column_letter(1 + len(bed_order))
    add_table(ws, f"A{header_row}:{second_end_letter}{data_end}", "BedroomPivotTable")
    ws.conditional_formatting.add(
        f"B{data_start}:{second_end_letter}{data_end}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 25
    for col in range(2, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.print_area = f"A1:{end_letter}{data_end}"


def build_raw_data(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw Data")
    set_common_sheet_style(ws)
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    style_header_row(ws, 1, 1, len(headers))
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, None if pd.isna(value) else value)
    last_row = len(df) + 1
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(len(headers))
    add_table(ws, f"A1:{last_col_letter}{last_row}", "RawRentDataTable")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    for col_idx, header in enumerate(headers, 1):
        width = 14
        if header in {"station", "rtb_area", "bedrooms", "property_type"}:
            width = 26
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_name in ("avg_rent", "dublin_avg"):
        col_idx = headers.index(col_name) + 1
        for row_idx in range(2, last_row + 1):
            ws.cell(row_idx, col_idx).number_format = '€#,##0.00'
    ws.print_title_rows = "1:1"


def build_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology")
    set_common_sheet_style(ws)
    style_title(ws, "A1:H1", "Methodology and Data Notes")
    sections = [
        ("Purpose", "Compare registered rents near all 24 Luas Green Line stops and test whether a longer commute is associated with lower rent."),
        ("Main comparison", "2025, 1 to 2 bed, all property types. This is the only bedroom-size category published for every RTB area mapped to the line."),
        ("RTB source", "https://data.cso.ie/table/RIA02"),
        ("Luas source", "https://data.tii.ie/Datasets/Luas/StopLocations/"),
        ("Rent definition", "Average monthly rent recorded for registered tenancies. It is not current asking rent from property listings."),
        ("Station mapping", "Each stop inherits the closest suitable RTB reporting area. Adjacent stops can therefore share the same rent observation."),
        ("Suppressed values", "Unpublished RTB cells are treated as missing observations, not zero rent."),
        ("Gradient", "Simple least-squares association between scheduled commute minutes and average rent. It is descriptive and not causal."),
        ("Limitations", "The analysis does not control for floor area, exact walking distance, building quality, tenancy start date or neighbourhood amenities."),
        ("Refresh", "Run python src/data_loader.py, python src/sql_utils.py --build, python src/export_powerbi.py, then python src/export_excel.py."),
        ("Pivot note", "The Pivot Summary sheet is generated from Raw Data as a reproducible pivot-style report. On Windows, select Raw Data and use Insert → PivotTable for an interactive native PivotTable."),
    ]
    ws.cell(3, 1, "Topic")
    ws.cell(3, 2, "Details")
    style_header_row(ws, 3, 1, 2)
    for row_idx, (topic, details) in enumerate(sections, 4):
        ws.cell(row_idx, 1, topic)
        ws.cell(row_idx, 2, details)
        ws.cell(row_idx, 1).font = Font(name="Aptos", bold=True, color=DARK)
        ws.cell(row_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        if details.startswith("https://"):
            ws.cell(row_idx, 2).hyperlink = details
            ws.cell(row_idx, 2).font = Font(name="Aptos", color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 34
    add_table(ws, f"A3:B{3 + len(sections)}", "MethodologyTable")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 105
    ws.freeze_panes = "A4"
    ws.print_area = f"A1:B{3 + len(sections)}"


def build_executive_summary(wb: Workbook, station_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Executive Summary", 0)
    set_common_sheet_style(ws)
    ws.sheet_view.showGridLines = False
    style_title(ws, "A1:L2", "Luas Green Line Rent Analysis")
    style_subtitle(
        ws,
        "A3:L3",
        "Executive report | RTB registered-tenancy rents | 24 stops | 2008–2025",
    )
    ws.merge_cells("A4:L4")
    ws["A4"] = "Selected view: 2025 · 1 to 2 bed · All property types"
    ws["A4"].font = Font(name="Aptos", size=10, color=DARK)
    ws["A4"].alignment = Alignment(horizontal="right")

    cards = [
        ("A6:C6", "A7:C9", "City-centre rent", "='Station Analysis'!D5", '€#,##0'),
        ("D6:F6", "D7:F9", "Cheapest rent", "=MIN('Station Analysis'!D5:D28)", '€#,##0'),
        ("G6:I6", "G7:I9", "Dearest rent", "=MAX('Station Analysis'!D5:D28)", '€#,##0'),
        ("J6:L6", "J7:L9", "Rent gradient", "=SLOPE('Station Analysis'!D5:D28,'Station Analysis'!C5:C28)", '€0.0'),
    ]
    for label_range, value_range, label, formula, fmt in cards:
        ws.merge_cells(label_range)
        ws.merge_cells(value_range)
        label_cell = ws[label_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        label_cell.value = label
        value_cell.value = formula
        label_cell.font = Font(name="Aptos", size=11, bold=True, color=WHITE)
        label_cell.fill = PatternFill("solid", fgColor=GREEN)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(name="Aptos Display", size=23, bold=True, color=DARK)
        value_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = fmt
        for row in ws[label_range]:
            for cell in row:
                cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
        for row in ws[value_range]:
            for cell in row:
                cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

    ws.merge_cells("D10:F10")
    ws["D10"] = "=INDEX('Station Analysis'!B5:B28,MATCH(MIN('Station Analysis'!D5:D28),'Station Analysis'!D5:D28,0))"
    ws.merge_cells("G10:I10")
    ws["G10"] = "=INDEX('Station Analysis'!B5:B28,MATCH(MAX('Station Analysis'!D5:D28),'Station Analysis'!D5:D28,0))"
    ws.merge_cells("J10:L10")
    ws["J10"] = "€ per extra commute minute"
    for cell in (ws["D10"], ws["G10"], ws["J10"]):
        cell.font = Font(name="Aptos", size=10, italic=True, color=DARK)
        cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A12:L12")
    ws["A12"] = "Decision-ready findings"
    ws["A12"].font = Font(name="Aptos Display", size=16, bold=True, color=DARK)
    findings = [
        "1. Commute time barely explains rent: the 2025 slope is about -€1.44 per minute and R² is only 0.01.",
        "2. Ranelagh combines a seven-minute commute with average rent of about €1,784 — €335 below the city-centre benchmark.",
        "3. The largest outward price jump is Cowper → Milltown: approximately +€396 per month for the same bedroom category.",
        "4. The corridor does not follow a simple 'farther out = cheaper' pattern; local area differences dominate.",
    ]
    for row_idx, finding in enumerate(findings, 13):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
        cell = ws.cell(row_idx, 1, finding)
        cell.font = Font(name="Aptos", size=11, color=BLACK)
        cell.fill = PatternFill("solid", fgColor=WHITE if row_idx % 2 else LIGHT_GREY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)
        ws.row_dimensions[row_idx].height = 27

    station_ws = wb["Station Analysis"]
    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "2025 Rent by Station"
    bar.x_axis.title = "Average monthly rent (€)"
    bar.height = 9
    bar.width = 14
    bar.add_data(Reference(station_ws, min_col=4, min_row=4, max_row=28), titles_from_data=True)
    bar.set_categories(Reference(station_ws, min_col=2, min_row=5, max_row=28))
    bar.legend = None
    ws.add_chart(bar, "A19")

    scatter = ScatterChart()
    scatter.title = "Rent vs Commute Time"
    scatter.x_axis.title = "Commute to city centre (min)"
    scatter.y_axis.title = "Average monthly rent (€)"
    scatter.height = 9
    scatter.width = 14
    scatter.scatterStyle = "marker"
    scatter.x_axis.scaling.min = 0
    scatter.x_axis.scaling.max = 45
    scatter.y_axis.scaling.min = 1500
    scatter.y_axis.scaling.max = 2500
    xvalues = Reference(station_ws, min_col=3, min_row=5, max_row=28)
    yvalues = Reference(station_ws, min_col=4, min_row=5, max_row=28)
    series = Series(yvalues, xvalues, title="Stations")
    series.marker.symbol = "circle"
    series.marker.size = 7
    series.marker.graphicalProperties.solidFill = GREEN
    series.marker.graphicalProperties.line.solidFill = DARK
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)
    ws.add_chart(scatter, "G19")

    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 12
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[6].height = 25
    for row in range(7, 10):
        ws.row_dimensions[row].height = 22
    ws.print_area = "A1:L37"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_workbook(output: Path) -> None:
    df = load_data()
    station_df = latest_station_slice(df)
    trend = yearly_trend(df)

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "Enzhe Shen"
    wb.properties.title = "Luas Green Line Rent Analysis — Executive Report"
    wb.properties.subject = "RTB rent analysis by Luas Green Line station"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    build_station_analysis(wb, station_df)
    build_trends(wb, trend)
    build_pivot_summary(wb, df)
    build_raw_data(wb, df)
    build_methodology(wb)
    build_executive_summary(wb, station_df)

    wb.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"wrote {output} ({output.stat().st_size / 1e6:.1f} MB)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    build_workbook(args.output)


if __name__ == "__main__":
    main()
